# -*- coding: utf-8 -*-
"""
This module offers a set of operations that user can modify their Excel files.
"""
import pathlib
import warnings
from os import remove
from pathlib import Path
from typing import List, Optional, Tuple, TypedDict, Union

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.cell.read_only import EmptyCell
from urllib3 import disable_warnings

from .excel_definition import ExcelDefinition, ExcelDefinitionColumn
from .sprint_schedule import SprintScheduleStore
from .story import Story, StoryFactory
from .utils import is_absolute_path_valid, standardize_column_name

__all__ = ["read_excel_file", "output_to_excel_file"]

# Currently, the openpyxl package will report an obsolete warning.
warnings.simplefilter(action="ignore", category=UserWarning)
# Disable the HTTPS certificate verification warning.
disable_warnings()


def read_excel_file(
    file: Union[str, Path],
    excel_definition: ExcelDefinition,
    sprint_schedule: SprintScheduleStore,
) -> Tuple[List[str], List[Story]]:
    """
    Read and parse the Excel file

    parm file:
        The Excel file that you want to read

    parm excel_definition:
        The Excel column definition which is imported
        from the :py:class:`ExcelDefinition`

    parm sprint_schedule:
        The priority mapping for the :py:class:`Milestone` object.

    :return:
        A :py:class:`tuple` object which contains a list of column
        name and a list of :py:class:`Story`.
    """
    work_book: Optional[Workbook] = None
    try:
        if not is_absolute_path_valid(file):
            raise FileNotFoundError(
                f"""Please make sure the input excel file exist and
                the path should be absolute. File: {file}."""
            )

        with open(str(file), mode="rb") as raw_file:
            work_book = openpyxl.load_workbook(
                raw_file,
                read_only=True,
                keep_vba=False,
                data_only=True,
                keep_links=True,
            )

            if work_book.active is None or (
                not isinstance(work_book.active, Worksheet)
                and not isinstance(work_book.active, ReadOnlyWorksheet)
            ):
                raise ValueError("The input excel file doesn't contain any sheets.")

            sheet: Union[Worksheet, ReadOnlyWorksheet] = work_book.active

            column_count = sheet.max_column
            if column_count is None:
                column_count = excel_definition.max_column_index

            if sheet.max_row is not None and sheet.max_row < 2:
                return [], []

            actual_excel_columns: List[str] = []
            actual_column_names: List[str] = []
            for column_index in range(1, column_count + 1):
                column_value = str(
                    sheet.cell(row=1, column=column_index).value  # type: ignore[misc]
                )
                if column_value is None or not str(column_value):
                    raise ValueError("The input excel file has invalid/empty column.")
                actual_excel_columns.append(column_value)
                if standardize_column_name(column_value) in actual_column_names:
                    raise ValueError(
                        f"""The input excel file has duplicate column.
Column name: {column_value}."""
                    )
                actual_column_names.append(standardize_column_name(column_value))

            stories: List[Story] = []

            defined_excel_columns = excel_definition.get_columns()

            class _DefinedColumn(TypedDict):
                finded: bool
                name: str
                self: ExcelDefinitionColumn

            defined_columns: List[_DefinedColumn] = [
                {
                    "finded": False,
                    "name": standardize_column_name(i["name"]),
                    "self": i,
                }
                for i in defined_excel_columns
            ]
            story_factory = StoryFactory(defined_excel_columns)

            for row in sheet.iter_rows(  # type: ignore[misc]
                min_row=2,
                max_row=sheet.max_row,
                min_col=1,
                max_col=len(actual_excel_columns),
            ):
                if __should_skip(row):
                    continue

                story: Story = story_factory.create_story()
                story.excel_row_index = __extract_row_number(row)
                actual_column_values: List[Cell] = list(row)

                for actual_column_index, actual_column_name in enumerate(
                    actual_column_names
                ):
                    if not actual_column_name:
                        continue
                    finded = False
                    for finded_defined_excel_column in defined_columns:
                        if finded_defined_excel_column[
                            "name"
                        ] == standardize_column_name(actual_column_name):
                            finded_defined_excel_column["finded"] = True
                            story.set_value(
                                finded_defined_excel_column["self"]["type"],
                                actual_column_name,
                                actual_column_values[actual_column_index].value,
                            )
                            finded = True
                            break
                    if not finded:
                        story.set_value(
                            str,
                            actual_column_name,
                            actual_column_values[actual_column_index].value,
                        )
                not_finded_defined_excel_column_names = [
                    i["self"]["name"] for i in defined_columns if i["finded"] is False
                ]
                if not_finded_defined_excel_column_names:
                    missing_msg = "\n".join(
                        [
                            f"{index + 1}.{name}"
                            for index, name in enumerate(
                                not_finded_defined_excel_column_names
                            )
                        ]
                    )
                    raise ValueError(f"Following columns are missing:\n{missing_msg}.")

                for item in defined_columns:
                    item["finded"] = False

                story.calc_sprint_schedule(sprint_schedule)
                stories.append(story)
        raw_file.close()
    finally:
        if work_book:
            work_book.close()
    return actual_excel_columns, stories


def __should_skip(row: Tuple[Cell, ...]) -> bool:
    is_all_cell_empty = True
    for cell in row:
        if (
            cell
            and not isinstance(cell, EmptyCell)
            and cell.value is not None
            and str(cell.value).strip()
        ):
            is_all_cell_empty = False
    return is_all_cell_empty


def __extract_row_number(row: Tuple[Cell, ...]) -> int:
    for cell in row:
        if cell and not isinstance(cell, EmptyCell):
            return int(cell.row)
    return 0


def output_to_excel_file(
    file: Union[str, Path],
    stories: "List[Story]",
    excel_column_names: List[str],
    over_write: bool = True,
):
    """
    Generate Excel file

    parm file:
        Output Excel file name including the path

    parm stories:
        A list of :py:class:`Story` which need to be written to the Excel

    parm excel_column_names:
        Excel columns
        the :py:class:`ExcelDefinition`

    parm over_write:
        Whether the exist output file will be over-write.
    """
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError("The output file path is invalid.")

    if pathlib.Path(file).exists():
        if over_write is True:
            try:
                remove(file)
            except PermissionError as e:
                raise FileExistsError(
                    f"The exist excel file: {file} cannot be removed. {e.args[0]}"
                ) from e
        else:
            raise FileExistsError(f"The output excel file: {file} is already exist.")

    work_book = openpyxl.Workbook(write_only=False)

    if work_book.active is None or (
        not isinstance(work_book.active, Worksheet)
        and not isinstance(work_book.active, Worksheet)
    ):
        work_book.close()
        raise ValueError("The output excel file cannot be generated.")

    sheet: Worksheet = work_book.active

    for column_index, column_name in enumerate(excel_column_names):
        cell = sheet.cell(row=1, column=column_index + 1)
        # There are three kinds of Cells. Only the Cell has the value attribute.
        if hasattr(cell, "value"):
            setattr(cell, "value", column_name)

    if stories:
        for row_index, story in enumerate(stories):
            for column_index, column_name in enumerate(excel_column_names):
                cell = sheet.cell(row=row_index + 2, column=column_index + 1)
                if hasattr(cell, "value"):
                    setattr(cell, "value", story.format_value(column_name))

    work_book.save(str(file))
    work_book.close()
